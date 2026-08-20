from openpyxl import load_workbook

INPUT_FILE = "input.xlsx"
OUTPUT_FILE = "output.xlsx"


def main():
    # Open Excel file
    workbook = load_workbook(INPUT_FILE)

    # Select worksheet
    sheet = workbook.active

    # Example: read data
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)

    # Example: add a new column
    sheet["D1"] = "Total"

    for row in range(2, sheet.max_row + 1):
        sheet[f"D{row}"] = f"=B{row}*C{row}"

    # Save result
    workbook.save(OUTPUT_FILE)

    print(f"Done! Saved as {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
