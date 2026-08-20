from openpyxl import Workbook
from random import randint, choice
from datetime import date, timedelta

OUTPUT_FILE = "input.xlsx"

# Create workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Sales Data"

# Headers
headers = [
    "Order ID",
    "Date",
    "Customer",
    "Product",
    "Category",
    "Quantity",
    "Price",
    "Total",
    "City",
]

sheet.append(headers)

# Sample data
customers = [
    "Rahul Sharma",
    "Amit Kumar",
    "Priya Singh",
    "Neha Verma",
    "Rohit Gupta",
    "Anjali Patel",
    "Vikas Jain",
    "Pooja Mehta",
]

products = [
    ("Laptop", "Electronics", 55000),
    ("Mouse", "Electronics", 800),
    ("Keyboard", "Electronics", 1500),
    ("Monitor", "Electronics", 12000),
    ("Chair", "Furniture", 7500),
    ("Desk", "Furniture", 10000),
    ("Notebook", "Stationery", 100),
    ("Pen", "Stationery", 30),
    ("Bag", "Accessories", 1200),
    ("Headphones", "Electronics", 2500),
]

cities = [
    "Bhopal",
    "Indore",
    "Delhi",
    "Mumbai",
    "Pune",
    "Jaipur",
    "Lucknow",
    "Ahmedabad",
]

start_date = date(2025, 1, 1)

# Generate 2000 rows
for i in range(1, 2001):
    customer = choice(customers)
    product, category, price = choice(products)
    quantity = randint(1, 10)
    order_date = start_date + timedelta(days=randint(0, 364))

    total = quantity * price

    sheet.append([
        i,
        order_date,
        customer,
        product,
        category,
        quantity,
        price,
        total,
        choice(cities),
    ])

# Basic formatting
for cell in sheet[1]:
    cell.font = cell.font.copy(bold=True)

# Date format
for row in range(2, sheet.max_row + 1):
    sheet[f"B{row}"].number_format = "DD-MM-YYYY"
    sheet[f"G{row}"].number_format = "₹#,##0"
    sheet[f"H{row}"].number_format = "₹#,##0"

# Make columns wider
widths = {
    "A": 12,
    "B": 15,
    "C": 20,
    "D": 18,
    "E": 18,
    "F": 12,
    "G": 15,
    "H": 15,
    "I": 15,
}

for column, width in widths.items():
    sheet.column_dimensions[column].width = width

# Save file
workbook.save(OUTPUT_FILE)

print(f"Excel file created successfully: {OUTPUT_FILE}")
print(f"Total data rows: {sheet.max_row - 1}")
